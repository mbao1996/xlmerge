#coding=gbk
from openpyxl import load_workbook
import os, time
from pandas.core.dtypes.inference import is_number

fn_config = "c:\\PythonWork\\test\\config.xlsx"
work_catalog = "c:\\PythonWork\\test\\"
fn = work_catalog + "demo.xlsx"
sheet = 'Sheet1'
opr = []

row_start = 3
col_opr = 7
#col_chg = [8,9,10]

def is_number(variate):
    flag = False
    if isinstance(variate,int):
        flag = True
    elif isinstance(variate,float):
        flag = True
    else:
        flag = False
    return(flag)
def add_str_to_list(list, str):
    str_exist = False
    for i in range(len(list)):
        if( list[i] == str ):
            str_exist = True
            break
    if( str_exist == False ):
        list.append(str)
def open_fn(fn, wr):
    try:
        if( wr == True ):
            wb = load_workbook(fn,keep_vba=True)
        else:
            wb = load_workbook(fn)
    except Exception as e:
        print(str(e))
        os._exit(0)
    return(wb)
def save_fn(wb, fn):
    try:
        wb.save(fn)
    except Exception as e:
        print(str(e))
        os._exit(0)
def read_config(fn):
    wb = open_fn(fn, False)
    ws = wb[sheet]
    rt = []
    col_chg = []
    if( ws.cell(1, 1).value == u"工作目录" ):
        data = ws.cell(1, 2).value
        work_catalog = ''.join(data.split())
    else:
        print('error in config "工作目录"')
        os._exit(0)
    if( ws.cell(2, 1).value == u"操作文件" ):
        data = ws.cell(2, 2).value
        fn = ''.join(data.split())
    else:
        print('error in config "操作文件"')
        os._exit(0)
    if( ws.cell(3, 1).value == u"起始行" ):
        row_start = ws.cell(3, 2).value
    else:
        print('error in config "起始行"')
        os._exit(0)
    if( ws.cell(4, 1).value == u"操作者列号" ):
        opr = ws.cell(4, 2).value
    else:
        print('error in config "操作者列号" ')
        os._exit(0)
    if( ws.cell(5, 1).value == u"待操作列" ):
        for i in range(2, ws.max_column+1):
            if( is_number(ws.cell(5, i).value)):
                col_chg.append(ws.cell(5, i).value)
            else:
                break
    else:
        print('error in config "待操作列" ')
        os._exit(0)
    rt.append(work_catalog)
    rt.append(fn)
    rt.append(row_start)
    rt.append(opr)
    rt.append(col_chg)
    return(rt)

cfg = read_config(fn_config)
work_catalog = cfg[0]
fn = work_catalog + cfg[1]
row_start = cfg[2]
col_opr = cfg[3]
col_chg = cfg[4]

wb = open_fn(fn, True)
ws = wb[sheet]
# get opr
for i in range(row_start, ws.max_row+1):
    op = ws.cell(i, col_opr).value
    op = ''.join(op.split())
    add_str_to_list(opr, op)
for k in range(len(opr)):   # every files to be merged
    print('-----------', opr[k])
    fn_div = fn.split('.')
    fn_rd = fn_div[0] + '_' + opr[k] + '.' + fn_div[1]

    wb_rd = open_fn(fn_rd, False)
    ws_rd = wb_rd[sheet]
    rd = None
    rd_sn = 0
    wr = None
    wr_sn = 0
    for i in range(row_start, ws_rd.max_row+1):     # 副文件的每一行
        if( ws_rd.cell(i, 1).value != None ):
            rd = ws_rd.cell(i, 1).value
            rd_sn = 0
        else:
            rd_sn += 1
        if( ws_rd.cell(i, col_opr).value == opr[k] ):   # 确认副文件的操作者
            for j in range(row_start, ws.max_row+1):    # 主文件中搜索
                if( ws.cell(j, 1).value != None ):
                    wr = ws.cell(j, 1).value
                    wr_sn = 0
                else:
                    wr_sn += 1
                if( ws.cell(j, col_opr).value == opr[k] ):   # 主文件中确认副文件的操作者
                    if( wr == rd and wr_sn == rd_sn ):      # 确认是同一行
                        for col in range(len(col_chg)):
                            ws.cell(j, col_chg[col]).value = ws_rd.cell(i, col_chg[col]).value
                        break

#save_fn(wb, fn)
#print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
time_flag = time.strftime("%Y%m%d_%H%M%S", time.localtime())
fn_div = fn.split('.')
fn_wr = fn_div[0] + '_' + time_flag + '.xlsm'
save_fn(wb, fn_wr)

print("work done")
